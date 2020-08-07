import ast
import copy
import json
from openpyxl import Workbook
import xlrd  # reading Excel
import csv
import cx_Oracle
import json

from .api import UsersAndGroups, User, Group, eprint

"""
Copyright 2018 ThoughtSpot
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

# -------------------------------------------------------------------------------------------------------------------

"""Classes to read and write users and groups."""


class UGXLSWriter:
    """
    Writes users and groups to an Excel spreadsheet.
    """

    def write(self, users_and_groups, filename):
        """
        Writes the content to the given file.
        :param users_and_groups:  The UsersAndGroups object to write.
        :type users_and_groups: UsersAndGroups
        :param filename:  Name of the file to write to.  No extension is expected and one will be added.
        :type filename: str
        """
        workbook = Workbook()
        workbook.remove(
            workbook.active
        )  # remove the default sheet since we'll be creating the ones we want.
        self._write_users(workbook, users_and_groups.get_users())
        self._write_groups(workbook, users_and_groups.get_groups())
        if not (filename.endswith("xls") or filename.endswith("xlsx")):
            filename += ".xlsx"

        workbook.save(filename)

    def _write_users(self, workbook, users):
        """
        Writes the users to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param users:  The list of groups to write.
        :type users: list of User
        :return:
        """
        ws = workbook.create_sheet(title="Users")
        self._write_header(
            ws,
            [
                "Name",
                "Password",
                "Display Name",
                "Email",
                "Groups",
                "Visibility"
            ],
        )
        cnt = 2  # start after header.
        for user in users:
            ws.cell(column=1, row=cnt, value=user.name)
            ws.cell(column=2, row=cnt, value=user.password)
            ws.cell(column=3, row=cnt, value=user.displayName)
            ws.cell(column=4, row=cnt, value=user.mail)
            ws.cell(column=5, row=cnt, value=json.dumps(user.groupNames))
            ws.cell(column=6, row=cnt, value=user.visibility)
            cnt += 1

    def _write_groups(self, workbook, groups):
        """
        Writes the groups to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param groups:  The list of groups to write.
        :type groups: list
        :return:
        """
        ws = workbook.create_sheet(title="Groups")
        self._write_header(
            ws,
            [
                "Name",
                "Display Name",
                "Description",
                "Groups",
                "Visibility",
                "Privileges",
            ],
        )
        cnt = 2  # start after header.
        for group in groups:
            ws.cell(column=1, row=cnt, value=group.name)
            ws.cell(column=2, row=cnt, value=group.displayName)
            ws.cell(column=3, row=cnt, value=group.description)
            ws.cell(column=4, row=cnt, value=json.dumps(group.groupNames))
            ws.cell(column=5, row=cnt, value=group.visibility)
            privileges = group.privileges if group.privileges else []
            ws.cell(column=6, row=cnt, value=json.dumps(privileges))
            cnt += 1

    @staticmethod
    def _write_header(worksheet, cols):
        """
        Writes the header for the given worksheet in row 1.
        :param worksheet:  Worksheet to write to.
        :param cols:  List of columns to write.
        """
        for ccnt in range(0, len(cols)):
            worksheet.cell(column=(ccnt + 1), row=1, value=cols[ccnt])


class UGXLSReader:
    """
    Reads user and group info from an Excel file that is formatted the same as the UGXLSWriter writes.
    """

    required_sheets = ["Users", "Groups"]
    required_columns = {
        "Users": [
            "Name",
            "Password",
            "Display Name",
            "Email",
            "Groups",
            "Visibility"
        ],
        "Groups": [
            "Name",
            "Display Name",
            "Description",
            "Groups",
            "Visibility"
        ],
    }

    def __init__(self):
        """
        Creates a new UGXLSReader
        """
        self.workbook = None
        self.indices = {}
        self.users_and_groups = UsersAndGroups()

    def read_from_excel(self, filepath):
        """
        Reads users and groups from the given file.
        :param filepath:  Path to the Excel file to read from.
        :type filepath: str
        :return: Returns the users and groups read from the Excel file.  The users and groups are not validated
        :rtype UsersAndGroups
        so that they can be modified prior to validation.
        """
        self.workbook = xlrd.open_workbook(filepath)
        if self._verify_file_format():
            self._get_column_indices()
            self._read_users_from_workbook()
            self._read_groups_from_workbook()
        return self.users_and_groups

    def _verify_file_format(self):
        """
        :return: True if the format of the workbook is valid.
        :rtype: bool
        """
        is_valid = True
        sheet_names = self.workbook.sheet_names()
        for required_sheet in UGXLSReader.required_sheets:
            if required_sheet not in sheet_names:
                eprint("Error:  missing sheet %s!" % required_sheet)
                is_valid = False
            else:
                sheet = self.workbook.sheet_by_name(required_sheet)
                header_row = sheet.row_values(rowx=0, start_colx=0)
                for required_column in UGXLSReader.required_columns[
                    required_sheet
                ]:
                    if required_column not in header_row:
                        eprint(
                            "Error:  missing column %s in sheet %s!"
                            % (required_column, required_sheet)
                        )
                        is_valid = False

        return is_valid

    def _get_column_indices(self):
        """
        Reads the sheets to get all of the column indices.  Assumes the format was already checked.
        """
        sheet_names = self.workbook.sheet_names()
        for sheet_name in sheet_names:
            if sheet_name in self.required_sheets:
                sheet = self.workbook.sheet_by_name(sheet_name)
                col_indices = {}
                ccnt = 0
                for col in sheet.row_values(rowx=0, start_colx=0):
                    col_indices[col] = ccnt
                    ccnt += 1
                self.indices[sheet_name] = col_indices

    def _read_users_from_workbook(self):
        """
        Reads all the users from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Users")
        indices = self.indices["Users"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # "Name", "Password", "Display Name", "Email", "Description", "Groups", "Visibility"
            username = row[indices["Name"]]
            password = row[indices["Password"]]
            display_name = row[indices["Display Name"]]
            email = row[indices["Email"]]
            groups = []
            if row[indices["Groups"]]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            visibility = row[indices["Visibility"]]

            try:
                user = User(
                    name=username,
                    password=password,
                    display_name=display_name,
                    mail=email,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one user per line.
                self.users_and_groups.add_user(
                    user, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint(f"Error reading user with name {username}")

    def _read_groups_from_workbook(self):
        """
        Reads all the groups from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Groups")
        indices = self.indices["Groups"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # Name", "Display Name", "Description", "Groups", "Visibility"
            group_name = row[indices["Name"]]
            display_name = row[indices["Display Name"]]
            description = row[indices["Description"]]
            visibility = row[indices["Visibility"]]

            groups = []
            if row[indices["Groups"]] and row[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            try:
                group = Group(
                    name=group_name,
                    display_name=display_name,
                    description=description,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one group per line.
                self.users_and_groups.add_group(
                    group, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except Exception:
                eprint("Error reading group with name %s" % group_name)

class UGCSVReader:
    """
    Reads users and groups from CSV. All users come from the user_csv file and
    groups are from the group_csv file.
    """
    DEFAULT_USER_FIELD_MAPPING = {
        "name": "Name",
        "display_name": "Display Name",
        "mail": "Email",
        "password": "Password",
        "group_names": "Groups",
        "visibility": "Visibility"
    }
    DEFAULT_GROUP_FIELD_MAPPING = {
        "name": "Name",
        "display_name": "Display Name",
        "description": "Description",
        "group_names": "Groups",
        "visibility": "Visibility",
        "privileges": "Privileges"
    }

    def __init__(self,
                 user_field_mapping=DEFAULT_USER_FIELD_MAPPING,
                 group_field_mapping=DEFAULT_GROUP_FIELD_MAPPING,
                 delimiter=","):
        """
        Creates a new CSV reader that can read based on the field mapping and delimiter.  While this class can
        cause groups to be created, the primary use is to have groups that will be.......??????????????????
        :param user_field_mapping: The mapping of columns to values for users.
        :type user_field_mapping: dict of str:str
        :param group_field_mapping: The mapping of columns to values for groups.
        :type group_field_mapping: dict of str:str
        :param delimiter: The delimiter to use.
        """
        self.user_field_mapping = copy.copy(user_field_mapping)
        self.group_field_mapping = copy.copy(group_field_mapping)
        self.delimiter = delimiter

        self.validate_fields()

    def validate_fields(self):
        """
        Verifies that the minimal required field mappings exist.  Raises a ValueError if not.
        :return: None
        :raises: ValueError
        """
        if "name" not in self.user_field_mapping.keys():
            raise ValueError("Missing mapping for 'name' for use with user CSV.")
        if "name" not in self.group_field_mapping.keys():
            raise ValueError("Missing mapping for 'name' for use with groups CSV.")

    def read_from_file(self, user_file, group_file=None):
        """
        Loads users and groups from the files.  If the group_file is not provided, the groups will be created from the
        user file with just the names.
        :param user_file: Path to the user file to read from.
        :type user_file: str
        :param group_file: Path to the group file to read from.
        :type group_file: str
        :return: Users and groups object.
        :rtype: UsersAndGroups
        """
        # initialize UsersAndGroups object to add User and Group objects to
        uag = UsersAndGroups()

        # Do minimal check on user CSV file, read, create User.

        # Saving the column name that "name" maps to since I use it again later
        user_name_column_name = self.user_field_mapping["name"]

        column_names = None

        with open(user_file, 'r') as uf:
            csv_reader = csv.reader(uf)
            csv_dict_reader = csv.DictReader(uf)
            firstline = 1
            for line in csv_dict_reader:
                #for the first line, check column names
                if firstline:
                    column_names = line.keys()
                    if user_name_column_name not in column_names:
                        raise ValueError("No column called '%s' in CSV" % user_name_column_name)
                # create User object

                #handle blanks in group_names column
                groups_field_raw = line[self.user_field_mapping["group_names"]]
                groups_field = "[]" if groups_field_raw == "" else groups_field_raw

                u = User(
                    name = line[user_name_column_name],
                    display_name = line[self.user_field_mapping["display_name"]],
                    mail = line[self.user_field_mapping["mail"]],
                    password = line[self.user_field_mapping["password"]],
                    group_names = ast.literal_eval(groups_field),# assumes valid list format, e.g. ["a", "b", ...]
                    visibility = line[self.user_field_mapping["visibility"]]
                    )
                #add User to UsersAndGroups object
                uag.add_user(u)
                firstline = 0


        # If there, do minimal check on group CSV file, read, create Group.

        # Saving the column name that "name" maps to since I use it again later
        group_name_column_name = self.group_field_mapping["name"]
        g_column_names = None

        if group_file is not None:
            with open(group_file, 'r') as gf:
                g_csv_reader = csv.reader(gf)
                firstline = 1
                g_csv_dict_reader = csv.DictReader(gf)
                for line in g_csv_dict_reader:
                    #for the first line, check column names
                    if firstline:
                        g_column_names = line.keys()
                        if group_name_column_name not in g_column_names:
                            raise ValueError("No column called '%s' in CSV" % group_name_column_name)
                        
                    if group_name_column_name not in g_column_names:
                        raise ValueError("No column called '%s' in CSV" % group_name_column_name)
                    # create Group object

                    #handle blanks in group_names column
                    g_groups_field_raw = line[self.group_field_mapping["group_names"]]
                    g_groups_field = "[]" if g_groups_field_raw == "" else g_groups_field_raw

                    g = Group(
                        name = line[group_name_column_name],
                        display_name = line[self.group_field_mapping["display_name"]],
                        description = line[self.group_field_mapping["description"]],
                        privileges = line[self.group_field_mapping["privileges"]],
                        group_names = ast.literal_eval(line[self.group_field_mapping["group_names"]]),# assumes valid list format, e.g. ["a", "b", ...]
                        visibility = line[self.group_field_mapping["visibility"]]
                        )
                    #add User to UsersAndGroups object
                    uag.add_group(g)
                    firstline = 0
        return uag



class UGOracleReader:
    """
    Reads users and groups from Oracle. 
    """
    DEFAULT_USER_FIELD_MAPPING = {
        "name": "Name",
        "display_name": "Display Name",
        "mail": "Email",
        "password": "Password",
        "group_names": "Groups",
        "group_names2": "Groups2",
        "visibility": "Visibility"
    }
    DEFAULT_GROUP_FIELD_MAPPING = {
        "name": "Name",
        "display_name": "Display Name",
        "description": "Description",
        "group_names": "Groups",
        "group_names2": "Groups2",
        "visibility": "Visibility",
        "privileges": "Privileges"
    }

    def __init__(self,
                 user_field_mapping=DEFAULT_USER_FIELD_MAPPING,
                 group_field_mapping=DEFAULT_GROUP_FIELD_MAPPING):
        """
        Creates a new Oracle reader.
        :param user_field_mapping: The mapping of columns to values for users.
        :type user_field_mapping: dict of str:str
        :param group_field_mapping: The mapping of columns to values for groups.
        :type group_field_mapping: dict of str:str
        """
        self.user_field_mapping = copy.copy(user_field_mapping)
        self.group_field_mapping = copy.copy(group_field_mapping)

        self.validate_fields()

    def validate_fields(self):
        """
        Verifies that the minimal required field mappings exist.  Raises a ValueError if not.
        :return: None
        :raises: ValueError
        """
        if "name" not in self.user_field_mapping.keys():
            raise ValueError("Missing mapping for 'name'.")
        if "name" not in self.group_field_mapping.keys():
            raise ValueError("Missing mapping for 'name'.")

    def read_from_oracle(self, oracle_u, oracle_pw, oracle_dsn, oracle_config, user_sql, group_sql, archive_dir, current_timestamp): # group_sql doesn't actually do anything with group_sql yet.
        """
        Loads users and groups from Oracle.  If the group_sql is not provided, the groups will be created from the
        user file with just the names.
        :param user_sql: Path to the user query SQL file.
        :type user_sql: str
        :param group_sql: Path to the group query SQL file.
        :type group_sql: str
        :return: Users and groups object.
        :rtype: UsersAndGroups
        """
        # check archive_dir (for achiving query results)
        if not archive_dir.endswith('/'):
            archive_dir += '/'

        # initialize UsersAndGroups object to add User and Group objects to
        uag = UsersAndGroups()

        # Read in Oracle connection config file, SQL file(s), run query, do minimal check on result, and create User.

        # Saving the column name that "name" maps to since I use it again later
        user_name_column_name = self.user_field_mapping["name"]

        if oracle_u and oracle_pw and oracle_dsn:
            connection = cx_Oracle.connect(oracle_u, oracle_pw, oracle_dsn) # If this causes error, try setting $TNS_ADMIN to the dir containing tnsnames.ora
        else:
            with open(oracle_config) as json_file:
                connect_data = json.load(json_file)

            user = connect_data["user"]
            password = connect_data["password"]
            dsn_dict = connect_data["dsn"]
            host = dsn_dict["host"]
            port = dsn_dict["port"]
            service_name = dsn_dict["service_name"]
            dsn = cx_Oracle.makedsn(host=host, port=port, service_name=service_name)
            # Connect
            connection = cx_Oracle.connect(user=user, password=password, dsn=dsn)
        # Query
        cursor = connection.cursor()
        cursor.execute("SET TRANSACTION READ ONLY")

        with open(user_sql) as sql_f:
            sql = sql_f.read()

        cursor.execute(sql)

        column_names = [col[0] for col in cursor.description]
        if user_name_column_name not in column_names:
            raise ValueError("No column called '%s' in query results" % user_name_column_name)
        query_results = cursor.fetchall() # a list

        # Create Users and also add to archive file

        user_archive_filename = '{0}users_to_sync_from_oracle{1}.csv'.format(archive_dir, current_timestamp)

        with open(user_archive_filename, 'w') as user_archive_file:
            user_writer = csv.DictWriter(user_archive_file, fieldnames=column_names)
            user_writer.writeheader()

            for tupl in query_results:
                line = {} # TODO maybe change name to line_dict
                for i in range(0, len(column_names)):
                    line.update({column_names[i]: tupl[i]})
                user_writer.writerow(line)

                groups_field_raw = line[self.user_field_mapping["group_names"]]
                groups_field = "[]" if groups_field_raw == "" else groups_field_raw

                if self.user_field_mapping["group_names2"] in line.keys():
                    groups_field2_raw = line[self.user_field_mapping["group_names2"]]
                else:
                    groups_field2_raw = "[]"
                groups_field2 = "[]" if groups_field_raw == "" else groups_field_raw # even if the column is there, it could be blank

                u = User(
                    name = line[user_name_column_name],
                    display_name = line[self.user_field_mapping["display_name"]],
                    mail = line[self.user_field_mapping["mail"]],
                    password = line[self.user_field_mapping["password"]],
                    group_names = ast.literal_eval(groups_field) + ast.literal_eval(groups_field2),# assumes valid list format, e.g. ["a", "b", ...]
                    visibility = line[self.user_field_mapping["visibility"]]
                    )
                #add User to UsersAndGroups object
                uag.add_user(u)




        # TODO If present, run group_sql query, do minimal checks, and create Groups from results.
        cursor.close()

        return uag

